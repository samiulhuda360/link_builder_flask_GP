import os
import time
import csv
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash, url_for, send_from_directory
from flask import stream_with_context, Response
from services import (get_all_sitenames, get_url_data_from_db, save_matched_to_excel, process_site, store_posted_url, extract_domain, delete_site_and_links, fetch_site_details, test_post_to_wordpress, delete_from_wordpress, find_post_id_by_url)
from flask_socketio import SocketIO, emit
import json
from functools import wraps
import openpyxl
from utils import get_api_keys
import pandas as pd
import sqlite3
from werkzeug.utils import secure_filename
import subprocess
import glob
from operator import itemgetter
from dotenv import load_dotenv
from random import randrange
from urllib.parse import urlparse

load_dotenv()

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"  # or "None" if necessary
app.config['SESSION_COOKIE_SECURE'] = False

app.secret_key = 'sdfadfasdfasdfasdfasdf'

uploaded_filename = None
Exact_MATCH = False
SKIP_COM_AU = False
ONLY_COM_AU = False
NO_BODY_IMAGE = False
SKIP_USED_DOMAINS = False



def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)

    return decorated_function
@app.route('/logs', methods=['GET'])
@login_required
def view_logs():
    # Authentication and authorization checks here
    cmd = ["/usr/bin/journalctl", "-u", "flask_app", "-n", "300"]

    try:
        log_data = subprocess.check_output(cmd).decode('utf-8')
        # error_logs = [line for line in log_data.splitlines() if "ERROR" in line.upper()]
        formatted_logs = "<br>".join(log_data.splitlines())
    except subprocess.CalledProcessError as e:
        return f"Error executing command: {str(e)}"

    return render_template("logs.html", log_data=formatted_logs)



#DataBase Operation
def get_link_list_from_db(host_url):
    # Establish a connection to the SQLite database
    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    # Query to retrieve all the links for a given site
    cursor.execute('''
    SELECT url 
    FROM links
    INNER JOIN sites ON links.site_id = sites.site_id
    WHERE sitename = ?
    ''', (host_url,))

    links = [row[0] for row in cursor.fetchall()]
    conn.close()

    return links

@app.route('/save_api_config', methods=['POST'])
def save_api_config():
    openai_api = request.form.get('openaiapi')
    pexels_api = request.form.get('pexelsapi')

    con = sqlite3.connect('api_config.db')
    cur = con.cursor()

    # Check if the table already contains data
    cur.execute("SELECT COUNT(*) FROM api_keys")
    count = cur.fetchone()[0]

    if count == 0:
        # Insert new APIs
        cur.execute("INSERT INTO api_keys (openai_api, pexels_api) VALUES (?, ?)", (openai_api, pexels_api))
    else:
        # Update existing APIs depending on which fields are filled
        if openai_api and pexels_api:
            cur.execute("UPDATE api_keys SET openai_api = ?, pexels_api = ? WHERE id = 1", (openai_api, pexels_api))
        elif openai_api:
            cur.execute("UPDATE api_keys SET openai_api = ? WHERE id = 1", (openai_api,))
        elif pexels_api:
            cur.execute("UPDATE api_keys SET pexels_api = ? WHERE id = 1", (pexels_api,))

    con.commit()
    con.close()

    flash("API configuration updated successfully!", "success")
    return redirect(url_for('config_manager'))

# Flask route to handle site deletion
@app.route('/delete-site', methods=['DELETE'])
def delete_site_route():
    sitename = request.args.get('sitename')
    if not sitename or sitename == "all":
        return jsonify({"error": "Invalid site name"}), 400

    delete_site_and_links(sitename)

    return jsonify({"message": "Site and associated links deleted successfully"}), 200


@app.route('/download_excel')
def download_excel():
    global uploaded_filename
    print(uploaded_filename)

    if not uploaded_filename:
        return jsonify({"error": "No current file running for download"}), 404

    excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_filename)

    # Send the file as a response with appropriate headers
    try:
        return send_file(excel_file_path, as_attachment=True, download_name=uploaded_filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/delete-all-excel-files', methods=['POST'])
@login_required
def delete_all_excel_files():
    excel_files = glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], "*.xlsx"))

    if not excel_files:
        return jsonify({"message": "No Excel files found to delete."}), 200

    for file_path in excel_files:
        os.remove(file_path)

    return jsonify({"message": "All Excel files deleted successfully"}), 200


@app.route('/get_files')
def get_files():
    # List all files in the UPLOAD_FOLDER
    files = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if
             os.path.isfile(os.path.join(app.config['UPLOAD_FOLDER'], f))]

    # Get creation time for each file and sort by latest
    files = sorted(files, key=lambda x: os.path.getctime(os.path.join(app.config['UPLOAD_FOLDER'], x)), reverse=True)

    return jsonify(files)

@app.route('/download_excel_from_file')
def download_excel_from_file():
    filename = request.args.get('filename')
    if filename:
        return send_from_directory(directory=app.config['UPLOAD_FOLDER'], path=filename, as_attachment=True)
    return "File not found", 404

@app.route('/failed_csv')
def failed_site_download_excel():
    # Specify the file path where the Excel file is saved
    failed_csv_file_path = 'failed_urls.csv'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(failed_csv_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/uploaded_excel')
def uploaded_download_excel():
    # Specify the file path where the Excel file is saved
    uploaded_excel_file_path = 'uploads/uploaded_excel.xlsx'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(uploaded_excel_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500




@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()

        admin_username = os.environ.get('ADMIN_USERNAME')
        admin_password = os.environ.get('ADMIN_PASSWORD')

        if username == admin_username and password == admin_password:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            flash('Incorrect credentials.')
            return redirect(url_for('login'))
    return render_template('login.html')



@app.route('/logout')
def logout():
    session.pop('logged_in', None)  # Clear the logged_in flag from the session.
    flash('You were successfully logged out.')
    return redirect(url_for('login'))  # Redirect to login page or another page of your choice.




@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/config', methods=['GET', 'POST'])
@login_required
def config_manager():
    if request.method == 'POST':
        openai_api = request.form.get('openaiapi').strip()
        pexels_api = request.form.get('pexelsapi').strip()

        con = sqlite3.connect('api_config.db')
        cur = con.cursor()

        # Check if the table already contains data
        cur.execute("SELECT COUNT(*) FROM api_keys")
        count = cur.fetchone()[0]

        if count == 0:
            # Insert new APIs
            cur.execute("INSERT INTO api_keys (openai_api, pexels_api) VALUES (?, ?)", (openai_api, pexels_api))
        else:
            # Update existing APIs depending on which fields are filled
            if openai_api and pexels_api:
                cur.execute("UPDATE api_keys SET openai_api = ?, pexels_api = ? WHERE id = 1", (openai_api, pexels_api))
            elif openai_api:
                cur.execute("UPDATE api_keys SET openai_api = ? WHERE id = 1", (openai_api,))
            elif pexels_api:
                cur.execute("UPDATE api_keys SET pexels_api = ? WHERE id = 1", (pexels_api,))

        con.commit()
        con.close()

        flash("API configuration updated successfully!", "success")

    # Fetch the keys to display
    api_keys = get_api_keys() or {"openai_api": "", "pexels_api": ""}
    return render_template('configuration.html', openai_api=api_keys["openai_api"], pexels_api=api_keys["pexels_api"])


@app.route('/site-manager', methods=['GET'])
@login_required
def site_manager():
    sitename_filter = request.args.get('sitename_filter', None)  # Notice the default value is now None

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    sites_data = []

    if sitename_filter:
        if sitename_filter == 'all':
            cursor.execute('SELECT * FROM sites')
        else:
            cursor.execute('SELECT * FROM sites WHERE sitename = ?', (sitename_filter,))

        sites = cursor.fetchall()

        for site in sites:
            site_id, sitename, username, app_password = site
            cursor.execute('SELECT url FROM links WHERE site_id = ?', (site_id,))
            links = cursor.fetchall()
            sites_data.append({
                'site_id': site_id,
                'sitename': sitename,
                'username': username,
                'app_password': app_password,
                'links': [link[0] for link in links]
            })

    # Fetching all site names for the dropdown
    cursor.execute('SELECT DISTINCT sitename FROM sites')
    all_sitenames = [row[0] for row in cursor.fetchall()]

    conn.close()

    return render_template('site_manager.html', filtered_sites=sites_data, all_sitenames=all_sitenames)


@app.route('/restapi_test')
def test_page():
    sitename_filter = request.args.get('sitename_filter', None)
    print(f"Received sitename_filter: {sitename_filter}")

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    message = ""

    if sitename_filter:
        print(f"Filtering by sitename: {sitename_filter}")
        cursor.execute('SELECT * FROM sites WHERE sitename = ?', (sitename_filter,))
        site = cursor.fetchone()
        print(f"Database query result: {site}")

        if site:
            site_id, sitename, username, app_password = site
            print(f"Fetched site details - ID: {site_id}, Name: {sitename}, Username: {username}")

            response = test_post_to_wordpress(sitename.strip(), username.strip(), app_password.strip(), "This is a test post from the API.")
            print(f"Post to WordPress response: {response}")

            if response and response.status_code == 201:
                post_id = response.json().get('id')
                print(f"Post created with ID: {post_id}")

                delete_response = delete_from_wordpress(sitename, username, app_password, post_id)
                print(f"Delete post response: {delete_response}")

                if delete_response and delete_response.status_code == 200:
                    message = f"Content posted to {sitename} successfully and then deleted."
                else:
                    message = f"Content posted to {sitename} successfully, but failed to delete."
            else:
                message = f"Failed to post on {sitename} {response}."
        else:
            print(f"No site found with sitename: {sitename_filter}")

    cursor.execute('SELECT DISTINCT sitename FROM sites')
    all_sitenames = [row[0] for row in cursor.fetchall()]
    print(f"All sitenames: {all_sitenames}")

    conn.close()

    print(f"Message to display: {message}")
    return render_template('restapi_test.html', message=message, all_sitenames=all_sitenames)


# Route to handle stopping the test
@app.route('/stop_test', methods=['POST'])
@login_required
def stop_test():
    global test_running
    test_running = False
    return jsonify({"message": "Test stopping initiated"}), 200

@app.route('/apitest', methods=['GET'])
@login_required
def apitest():
    global test_running
    test_running = True
    site_details = fetch_site_details()
    failed_sites_excel_path = 'failed_sites.xlsx'

    # Initialize a new blank Excel file
    if os.path.exists(failed_sites_excel_path):
        os.remove(failed_sites_excel_path)
    pd.DataFrame(columns=['Failed Site URL', 'Error Code']).to_excel(failed_sites_excel_path, index=False)

    if site_details:
        for site in site_details:
            if not test_running:
                break  # Stop the testing if the global variable is False
            site_url, username, app_password = site
            content = "This is a test post from the API."
            time.sleep(randrange(3, 7))  # Simulate a delay
            try:
                success = test_post_to_wordpress(site_url.strip(), username.strip(), app_password.strip(), content)
                print(success.status_code)
                if success.status_code == 201:  # Assuming success is a response object
                    post_id = success.json().get('id')
                    delete_response = delete_from_wordpress(site_url, username, app_password, post_id)
                    if delete_response and delete_response.status_code == 200:
                        print(f"Post deleted successfully from {site_url}.")
                        message = f"Post published successfully to {site_url}."
                    else:
                        print(f"Failed to delete post from {site_url}.")
                        raise Exception("Failed to delete post")
                else:
                    raise Exception(f"Failed to publish post. Status code: {success.status_code}")
            except Exception as e:
                # Handle any kind of exception, log to Excel
                print(f"Exception for {site_url}: {str(e)}")
                error_code = str(e).split(":")[-1].strip() if "Status code" in str(e) else "N/A"
                existing_df = pd.read_excel(failed_sites_excel_path)
                new_df = pd.DataFrame([[site_url, error_code]], columns=['Failed Site URL', 'Error Code'])
                updated_df = pd.concat([existing_df, new_df], ignore_index=True)
                updated_df.to_excel(failed_sites_excel_path, index=False)
                message = f"Failed to publish post to {site_url}."

            # Update via socketio
            socketio.emit('apitest_update', {'message': message})

    # Completion message
    socketio.emit('apitest_complete', {'message': "API Test completed"})
    return jsonify({"message": "API Test completed"}), 200



@app.route('/upload-excel', methods=['POST'])
@login_required
def upload_excel_site_data():
    if 'excel_data' not in request.files:
        return "No file part", 400

    file = request.files['excel_data']

    if file.filename == '':
        return "No selected file", 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return "Invalid file type. Please upload an Excel file.", 400

    df = pd.read_excel(file, engine='openpyxl')

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    for _, row in df.iterrows():
        sitename, username, app_password = row['Sitename'], row['Username'], row['Application_Password']
        added_link = row['Added_Link'] if 'Added_Link' in row and not pd.isnull(row['Added_Link']) else None

        cursor.execute('SELECT site_id FROM sites WHERE sitename=?', (sitename,))
        site_id = cursor.fetchone()

        if site_id:  # if site exists
            cursor.execute('UPDATE sites SET username=?, app_password=? WHERE site_id=?',
                           (username, app_password, site_id[0]))
        else:  # if site doesn't exist
            cursor.execute('INSERT INTO sites (sitename, username, app_password) VALUES (?, ?, ?)',
                           (sitename, username, app_password))
            site_id = (cursor.lastrowid,)

        # Insert link if it doesn't exist for the site and the link is not None
        if added_link is not None:
            cursor.execute('SELECT url FROM links WHERE site_id=? AND url=?', (site_id[0], added_link))
            if not cursor.fetchone():
                cursor.execute('INSERT INTO links (site_id, url) VALUES (?, ?)', (site_id[0], added_link))

    conn.commit()
    conn.close()
    flash('Site Data Updated successfully!', 'success')
    return redirect(url_for('site_manager'))




def update_excel_with_live_link(file_path, row_index, live_url):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Assuming you want to write to column 'J=Live_link'
    sheet[f'J{row_index}'] = live_url

    wb.save(file_path)

@app.route('/stop_processing', methods=['POST'])
def stop_processing():
    global should_continue_processing
    should_continue_processing = False
    return jsonify({"message": "Processing will be stopped."}), 200


@app.route('/start_emit', methods=['POST'])
def start_emit():
    global should_continue_processing
    global USE_IMAGES
    should_continue_processing = True

    # Get the Excel file
    excel_file = request.files['excel_file']
    if not excel_file:
        return jsonify({"error": "No file uploaded."}), 400

    # Set USE_IMAGES based on checkbox value
    USE_IMAGES = 'use_images' in request.form
    Exact_MATCH = 'exact_match' in request.form
    SKIP_COM_AU = 'skip_au' in request.form
    ONLY_COM_AU = 'only_au' in request.form
    NO_BODY_IMAGE = 'no_body_image' in request.form
    print("NO BODY IMAGE:", NO_BODY_IMAGE  )
    used_domains = set()  # Step 1: Initialize the set
    SKIP_USED_DOMAINS = 'skip_used_domains' in request.form  # Step 2: Check the option from frontend

    global uploaded_filename
    original_filename = secure_filename(excel_file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], original_filename)
    excel_file.save(file_path)

    # Store the filename in the global variable
    uploaded_filename = original_filename
    print(uploaded_filename)

    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Create an empty list to accumulate the data
    data_list = []
    sitenames = get_all_sitenames()
    num_sites = len(sitenames)

    rows = list(sheet.iter_rows(values_only=True))
    total_rows = len(rows)
    row_index = 0

    last_used_site_index = -1  # Start with -1 so that for the first row, it starts with 0.
    with open('failed_urls.csv', 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Failed URLs"])
    try:
        while row_index < total_rows and should_continue_processing:
            row = rows[row_index]

            if row_index == 0:
                row_index += 1
                continue
            if row[1] is None or row[2] is None:
                row_index += 1
                continue

        # asyncio.run(my_async_function())
            anchor, linking_url, embed_code, map_embed_title, name, address, phone, topic, live_link = row[1:10]


            if row[9] != None and row[9] != "Failed To Post":
                print("skipping the row")
                row_index += 1
                continue


            link_posted = False
            failed_post_count = 0  # Counter for failed posts.
            start_site_index = (last_used_site_index + 1) % num_sites  # Start from the next site.



            for offset in range(num_sites):
                site_index = (start_site_index + offset) % num_sites  # Wrap around using modulo.
                host_url = sitenames[site_index].strip()
                print("Posting to:", host_url)
                link_list = get_link_list_from_db(host_url)
                # Extract domains from link_list
                domain_list = [extract_domain(link) for link in link_list]
                # Inside the success condition where you've successfully posted the link:
                last_used_site_index = site_index  # Update the last used site index.

                # Step 3: Check against the set
                if SKIP_USED_DOMAINS and host_url in used_domains:
                    print(f"Skipping {host_url} because it has been used before.")
                    continue



                # Exact Match Check
                if Exact_MATCH:
                    if linking_url in link_list:
                        print(linking_url)
                        continue  # Skip to the next iteration

                # If Exact_MATCH is False, then it's Root Match
                elif extract_domain(linking_url) in domain_list:
                    print(linking_url)
                    print("Matched Root Domain inside")
                    continue  # Skip to the next iteration
                # Skip com.au and org.au Check
                if SKIP_COM_AU:
                    print("Checking URL:", host_url)
                    if host_url.endswith('com.au') or host_url.endswith('org.au'):
                        print("Skipping because it ends with com.au or org.au")
                        continue  # Skip to the next iteration

                # Only com.au and org.au Check
                if ONLY_COM_AU:
                    if not (host_url.endswith('com.au') or host_url.endswith('org.au')):
                        print("Skipping because it doesn't end with com.au or org.au")
                        continue  # Skip to the next iteration



                if not should_continue_processing:
                    break


                user_password_data = get_url_data_from_db(host_url)
                site_json = "https://" + host_url.strip() + "/wp-json/wp/v2"

                if user_password_data:
                    user = user_password_data.get('user')
                    password = user_password_data.get('password')

                    # for _ in range(5):
                    #     asyncio.run(my_async_function())
                    if not should_continue_processing:
                        break

                    # Testing START

                    if name is None:
                        name = ""
                    else:
                        name = name + "<br>"
                    if address is None:
                        address = ""
                    else:
                        address = address + "<br>"
                    if phone is None:
                        phone = ""

                    nap = str(name) + str(address) + str(phone) +"<br>"

                    live_url = process_site(site_json, user.strip(), password.strip(), topic, anchor, linking_url, embed_code,
                                            map_embed_title, nap, USE_IMAGES, NO_BODY_IMAGE)

                    if live_url == "Failed To Post":
                        with open('failed_urls.csv', 'a', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow([host_url])
                            print("Failed at:", host_url)
                        failed_post_count += 1
                        if failed_post_count < 5:
                            continue
                        else:
                            break
                            # Step 4: Add to the set
                    used_domains.add(host_url)
                    update_excel_with_live_link(file_path, row_index + 1, live_url) # Updating Excel

                    if live_url != "Failed To Post":
                        # Adding to Database
                        store_posted_url(host_url, linking_url) # Adding the posted link to database
                    else:
                        print("Not Posting")
                    # Testing END

                    data = {
                        'id': row_index,
                        'anchor': anchor,
                        'linking_url': linking_url,
                        'nap': nap,
                        'topic': topic,
                        'live_url': live_url,
                        "Host_site": host_url,
                    }
                    time.sleep(1)

                    data_list.append(data)
                    socketio.emit('update', {'data': json.dumps(data_list)})


                    link_posted = True
                    break

            if not link_posted:
                print("All Sites have this link")
                pass

            row_index += 1

        if not should_continue_processing:
            print("Process stopped")
            flash("Process stopped")
            socketio.emit('update', {"message": "Process stopped"})
            return jsonify({"message": "Processing was halted by the user."}), 200
    except Exception as e:
        # Log the error for debugging purposes
        print(f"An error occurred: {str(e)}")
        # Send an error message to the frontend
        socketio.emit('error', {'message': str(e)})

    flash("Processed successfully!")
    socketio.emit('update', {"message": "Processing Ended"})
    return jsonify({"message": "Processing Ended"}), 200


@app.route('/download_failed_sites')
def download_failed_sites():
    # Specify the file path where the Excel file is saved
    failed_csv_file_path = 'failed_sites.xlsx'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(failed_csv_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download_excel_template')
def download_excel_template():
    # Specify the file path where the Excel file is saved
    excel_template_file_path = 'excel_template_pbn.xlsx'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(excel_template_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    
@app.route('/post_delete', methods=['GET'])
def post_delete():
    return render_template('post_delete.html')

@socketio.on('delete_request')
def handle_delete_request(data):
    urls = data['urls']

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    for url in urls:
        try:
            # Parse the URL to extract the domain
            parsed_url = urlparse(url)
            domain = parsed_url.netloc
            print(f"Extracted domain: {domain}")  # Logging statement

            if domain:
                # Query the sites_data database to get the site details
                cursor.execute('SELECT * FROM sites WHERE sitename = ?', (domain,))
                site = cursor.fetchone()

                if site:
                    site_id, sitename, username, app_password = site
                    print(f"Retrieved site details: {sitename}, {username}, {app_password}")  # Logging statement

                    # Find the post ID using the find_post_id_by_url function
                    post_id = find_post_id_by_url(sitename, url, username, app_password)
                    print(f"Found Post ID: {post_id}")  # Logging statement

                    if post_id:
                        # Delete the post using the delete_from_wordpress function
                        response = delete_from_wordpress(sitename, username, app_password, post_id)
                        if response is not None and response.status_code == 200:
                            result = f"Post deleted successfully: {url}"
                        else:
                            result = f"Failed to delete post: {url}"
                    else:
                        result = f"Post not found for URL: {url}"
                else:
                    result = f"Site not found for URL: {url}"
            else:
                result = f"Invalid URL: {url}"
                print(result)
        except Exception as e:
            result = f"Error processing URL: {url} - {str(e)}"
        print(result)
        socketio.emit('delete_update', {'message': result})

    conn.close()
    socketio.emit('delete_complete', {'message': 'Delete operation completed.'})

if __name__ == '__main__':
    socketio.run(app, debug=True)



